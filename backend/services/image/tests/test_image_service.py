"""图片识别 Agent 单元测试：预处理、识别解析、LLM 整理、Excel 导出、SSE 编排。"""

from __future__ import annotations

import base64
import io
from pathlib import Path
from types import SimpleNamespace

import pytest

from backend.services.glm46v.client import GLM46VSettings
from backend.services.image.excel import (
    build_excel_filename,
    is_safe_excel_filename,
    write_recognition_excel,
)
from backend.services.image.models import (
    ImageRecognitionFailure,
    SheetRecognition,
    backfill_empty_slots,
)
from backend.services.image.preprocess import preprocess_image
from backend.services.image.recognition import (
    build_recognition_prompt,
    classify_failure_kind,
    merge_recognitions,
    recognize_image_segments,
    recognize_single_image,
)
from backend.services.image.structuring import _deterministic_rows, structure_rows
from backend.services.llm.credentials import LlmCredentials


def tiny_png_base64() -> str:
    return base64.b64encode(
        base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
        )
    ).decode("ascii")


def _make_png(width: int, height: int) -> str:
    from PIL import Image

    buffer = io.BytesIO()
    Image.new("RGB", (width, height), (220, 220, 220)).save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode("ascii")


class FakeClient:
    def __init__(self, content: str = "") -> None:
        self.content = content
        self.calls: list[str] = []

    @property
    def settings(self) -> GLM46VSettings:
        return GLM46VSettings(api_key="fake", endpoint="https://example.test")

    async def analyze_images(self, images, *, prompt: str, max_tokens: int = 6144):
        self.calls.append(prompt)
        if not self.content:
            raise RuntimeError("mock 识别失败")
        return {"model": "glm-4.6v-flash", "content": self.content, "usage": {}}


# ---------- preprocess ----------


def test_preprocess_enhances_and_upscales() -> None:
    image = preprocess_image(
        name="shelf.png",
        mime_type="image/png",
        data=_make_png(80, 60),
        max_image_mb=20,
    )
    assert image.mime_type == "image/jpeg"
    assert image.size_bytes > 0
    from PIL import Image

    decoded = Image.open(io.BytesIO(base64.b64decode(image.data)))
    assert decoded.width == 160  # 2x 放大
    assert decoded.height == 120
    assert image.name == "shelf.png"


def test_preprocess_rejects_invalid_data() -> None:
    from backend.services.glm46v.client import GLM46VError

    with pytest.raises(GLM46VError):
        preprocess_image(
            name="bad.png",
            mime_type="image/png",
            data="not-base64-!",
            max_image_mb=20,
        )


# ---------- recognition ----------


def test_build_recognition_prompt_contains_fixed_rules() -> None:
    prompt = build_recognition_prompt("shelf_a.jpg")
    assert "从上到下" in prompt
    assert "空位写“空”" in prompt
    assert "编号无法辨认" in prompt
    assert "叠放的每一张都要单独识别" in prompt
    assert "shelf_a.jpg" in prompt


@pytest.mark.asyncio
async def test_recognize_single_image_parses_grid_with_empty_and_uncertain() -> None:
    """整图路径：完整网格模板保留空位与无法辨认占位，位号不被压缩。"""

    content = (
        "第1层：第1位=325, 第2位=空, 第3位=编号无法辨认, 第4位=43, 第5位=268\n"
        "第2层：第1位=224, 第2位=89\n"
    )
    client = FakeClient(content)
    rows, summary, error = await recognize_single_image(
        image=SimpleNamespace(name="shelf_a.jpg"),
        client=client,
        prompt="识别",
    )
    assert error is None
    assert [(item.sheet_no, item.layer, item.position, item.note) for item in rows] == [
        ("325", 1, 1, ""),
        ("", 1, 2, "空货位"),
        ("", 1, 3, "编号无法辨认"),
        ("43", 1, 4, ""),
        ("268", 1, 5, ""),
        ("224", 2, 1, ""),
        ("89", 2, 2, ""),
    ]
    assert "1 个空货位" in summary
    assert "1 处无法辨认" in summary


@pytest.mark.asyncio
async def test_recognize_single_image_parses_grid_stack_rows() -> None:
    """整图路径：网格模板中的“第N层第M排”叠放行保留排号。"""

    content = (
        "第1层：第1位=003, 第2位=005\n"
        "第1层第2排：第1位=004\n"
    )
    client = FakeClient(content)
    rows, _summary, error = await recognize_single_image(
        image=SimpleNamespace(name="shelf_a.jpg"),
        client=client,
        prompt="识别",
    )
    assert error is None
    assert [(item.sheet_no, item.layer, item.stack, item.position) for item in rows] == [
        ("003", 1, 1, 1),
        ("005", 1, 1, 2),
        ("004", 1, 2, 1),
    ]


@pytest.mark.asyncio
async def test_recognize_single_image_trims_hallucinated_layer_width() -> None:
    """免费模型偶发“某层 16 位”虚构：明显超宽的层被截断到中位数宽度。"""

    content = (
        "第1层：第1位=1, 第2位=2, 第3位=3, 第4位=4, 第5位=5, 第6位=6, 第7位=7\n"
        "第2层：第1位=8, 第2位=9, 第3位=10, 第4位=11, 第5位=12, 第6位=13, 第7位=14\n"
        "第3层：第1位=15, 第2位=16, 第3位=17, 第4位=18, 第5位=19, 第6位=20, "
        "第7位=21, 第8位=22, 第9位=23, 第10位=24, 第11位=25, 第12位=26, "
        "第13位=27, 第14位=28, 第15位=29, 第16位=30\n"
    )
    client = FakeClient(content)
    rows, _summary, error = await recognize_single_image(
        image=SimpleNamespace(name="shelf_a.jpg"),
        client=client,
        prompt="识别",
    )
    assert error is None
    third_layer = [item for item in rows if item.layer == 3]
    assert max(item.position for item in third_layer) == 7


def test_backfill_empty_slots_backfills_gaps() -> None:
    """定位路径：绝对列号中间缺失的货位回填为“空货位”占位。"""

    rows = backfill_empty_slots(
        [
            SheetRecognition("002", layer=1, position=2, stack=1, source_image="a.jpg"),
            SheetRecognition("003", layer=1, position=3, stack=1, source_image="a.jpg"),
        ]
    )
    assert [(item.position, item.sheet_no, item.note) for item in rows] == [
        (2, "002", ""),
        (3, "003", ""),
        (1, "", "空货位"),
    ]


def test_parse_compact_output_handles_stacks_and_continuation() -> None:
    """紧凑网格格式：叠放排、续行、空位与无法辨认占位都能解析。"""

    from backend.services.image.recognition import _parse_compact_output

    content = (
        "第1层:1=325,2=空,3=编号无法辨认\n"
        ";2排:1=179,2=222\n"
        "第2层:1=62;2排:1=71,2=240"
    )
    rows = _parse_compact_output(content, source_image="shelf.jpg")
    assert [(item.sheet_no, item.layer, item.stack, item.position, item.note) for item in rows] == [
        ("325", 1, 1, 1, ""),
        ("", 1, 1, 2, "空货位"),
        ("", 1, 1, 3, "编号无法辨认"),
        ("179", 1, 2, 1, ""),
        ("222", 1, 2, 2, ""),
        ("62", 2, 1, 1, ""),
        ("71", 2, 2, 1, ""),
        ("240", 2, 2, 2, ""),
    ]


@pytest.mark.asyncio
async def test_recognize_image_segments_merges_layer_offsets() -> None:
    """分段识别：两段层号偏移合并，且请求带 max_tokens=1024、关闭 thinking。"""

    from PIL import Image as PILImage

    class FakeSegmentClient:
        def __init__(self, responses: list[str]) -> None:
            self.responses = list(responses)
            self.calls: list[tuple[int, int, bool]] = []

        @property
        def settings(self) -> GLM46VSettings:
            return GLM46VSettings(api_key="fake", endpoint="https://example.test")

        async def analyze_images(
            self,
            images,
            *,
            prompt: str,
            max_tokens: int = 6144,
            include_thinking: bool = True,
        ):
            self.calls.append((len(images), max_tokens, include_thinking))
            return {
                "model": "glm-4v-flash",
                "content": self.responses.pop(0),
                "usage": {},
            }

    client = FakeSegmentClient(
        [
            "第1层:1=003\n第2层:1=005",
            "第1层:1=008\n第2层:1=009;2排:1=010",
        ]
    )
    buffer = io.BytesIO()
    PILImage.new("RGB", (60, 300), (200, 200, 200)).save(buffer, format="PNG")
    image = SimpleNamespace(
        name="shelf.jpg",
        mime_type="image/png",
        data=base64.b64encode(buffer.getvalue()).decode("ascii"),
        size_bytes=len(buffer.getvalue()),
    )
    rows, summary, error = await recognize_image_segments(
        image=image,
        client=client,
        source_name="shelf.jpg",
    )
    assert error is None
    assert [(item.sheet_no, item.layer, item.stack, item.position) for item in rows] == [
        ("003", 1, 1, 1),
        ("005", 2, 1, 1),
        ("008", 3, 1, 1),
        ("009", 4, 1, 1),
        ("010", 4, 2, 1),
    ]
    assert client.calls == [(1, 1024, False), (1, 1024, False)]
    assert "第1段识别 2 个编号" in summary
    assert "第2段识别 3 个编号" in summary


@pytest.mark.asyncio
async def test_recognize_single_image_parses_pipe_rows() -> None:
    client = FakeClient("003|第1层|第2位|\n005|第2层|第3位|编号无法辨认")
    rows, summary, error = await recognize_single_image(
        image=SimpleNamespace(name="shelf_a.jpg"),
        client=client,
        prompt="识别",
    )
    assert error is None
    assert len(rows) == 2  # 005 标注无法辨认 → 保留为占位行
    assert rows[0].sheet_no == "003"
    assert rows[0].layer == 1
    assert rows[0].position == 2
    assert rows[0].source_image == "shelf_a.jpg"
    assert rows[1].sheet_no == ""
    assert rows[1].note == "编号无法辨认"
    assert rows[1].layer == 2
    assert rows[1].position == 3
    assert "识别到 1 个图纸编号" in summary


@pytest.mark.asyncio
async def test_recognize_single_image_parses_real_glm_output() -> None:
    """GLM-4.6V 真实输出：编号带方括号、多条记录挤在一行、空格分隔。"""

    content = (
        "[325]|[第1层|第1位| [89]|[第1层|第2位| [302]|[第1层|第3位| "
        "[386]|[第1层|第4位| [88]|[第1层|第5位| [734]|[第1层|第6位|"
    )
    client = FakeClient(content)
    rows, _summary, error = await recognize_single_image(
        image=SimpleNamespace(name="shelf_a.jpg"),
        client=client,
        prompt="识别",
    )
    assert error is None
    assert len(rows) == 6
    assert [item.sheet_no for item in rows] == ["325", "89", "302", "386", "88", "734"]
    assert rows[0].layer == 1
    assert rows[0].position == 1
    assert rows[-1].position == 6


@pytest.mark.asyncio
async def test_recognize_single_image_keeps_uncertain_placeholders() -> None:
    """整图路径：标注“编号无法辨认/不确定”的条目保留坐标占位。"""

    content = "005|第2层|第3位|编号无法辨认\n[008]|[第2层|第4位|"
    client = FakeClient(content)
    rows, _summary, error = await recognize_single_image(
        image=SimpleNamespace(name="shelf_b.jpg"),
        client=client,
        prompt="识别",
    )
    assert error is None
    assert len(rows) == 2
    assert rows[0].sheet_no == ""  # 005 不确定 → 占位
    assert rows[0].layer == 2
    assert rows[0].position == 3
    assert rows[0].note == "编号无法辨认"
    assert rows[1].sheet_no == "008"
    assert rows[1].note == ""


@pytest.mark.asyncio
async def test_recognize_single_image_parses_stack_rows() -> None:
    """整图路径：识别“第N排”叠放信息，叠放不再被拍平。"""

    content = (
        "003|第1层|第1排|第2位|\n"
        "004|第1层|第2排|第2位|\n"
        "005|第1层|第3位|"
    )
    client = FakeClient(content)
    rows, summary, error = await recognize_single_image(
        image=SimpleNamespace(name="shelf_a.jpg"),
        client=client,
        prompt="识别",
    )
    assert error is None
    assert [(item.sheet_no, item.layer, item.stack, item.position) for item in rows] == [
        ("003", 1, 1, 2),
        ("004", 1, 2, 2),
        ("005", 1, 1, 3),
    ]
    assert "识别到 3 个图纸编号" in summary


@pytest.mark.asyncio
async def test_recognize_single_image_keeps_uncertain_without_number() -> None:
    """整图路径：只有“第N层|第M位|编号无法辨认”也要保留占位。"""

    content = "第2层|第3位|编号无法辨认\n008|第1层|第4位|"
    client = FakeClient(content)
    rows, _summary, error = await recognize_single_image(
        image=SimpleNamespace(name="shelf_b.jpg"),
        client=client,
        prompt="识别",
    )
    assert error is None
    assert len(rows) == 2
    assert rows[0].sheet_no == ""
    assert rows[0].layer == 2
    assert rows[0].position == 3
    assert rows[0].note == "编号无法辨认"
    assert rows[1].sheet_no == "008"


@pytest.mark.asyncio
async def test_recognize_single_image_degrades_on_failure() -> None:
    client = FakeClient(content="")  # analyze_images 抛 RuntimeError
    rows, _summary, error = await recognize_single_image(
        image=SimpleNamespace(name="shelf_b.jpg"),
        client=client,
        prompt="识别",
    )
    assert rows == []
    assert error is not None
    assert "识别失败" in error


@pytest.mark.asyncio
async def test_recognize_single_image_handles_no_shelf() -> None:
    client = FakeClient("未检测到货架图纸")
    rows, summary, error = await recognize_single_image(
        image=SimpleNamespace(name="shelf_c.jpg"),
        client=client,
        prompt="识别",
    )
    assert rows == []
    assert error is None  # 未检测到图纸不算失败
    assert "未检测到货架图纸" in summary


def test_merge_recognitions_keeps_successes_and_failures() -> None:
    rows, failures, summaries = merge_recognitions(
        [
            (
                "a.jpg",
                [SheetRecognition("001", layer=1, position=1, stack=1, source_image="a.jpg")],
                None,
            ),
            ("b.jpg", [], "视觉识别失败：mock"),
            ("c.jpg", [], "智谱 API 请求失败（HTTP 429）：访问量过大，请稍后再试"),
        ]
    )
    assert len(rows) == 1
    assert len(failures) == 2
    assert failures[0].image_name == "b.jpg"
    assert failures[0].kind == "other"
    assert "mock" in failures[0].reason
    assert failures[1].kind == "rate_limited"  # 429 限流单独分类
    assert len(summaries) == 1  # 失败张不计入总结


def test_classify_failure_kind() -> None:
    assert classify_failure_kind("HTTP 429 访问量过大") == "rate_limited"
    assert classify_failure_kind("限流，请稍后再试") == "rate_limited"
    assert classify_failure_kind("图片无法解码") == "other"


# ---------- structuring ----------


def test_deterministic_rows_sorts_by_layer_then_position() -> None:
    rows = _deterministic_rows(
        [
            SheetRecognition("010", layer=10, position=1, stack=1, source_image="a.jpg"),
            SheetRecognition("002", layer=2, position=1, stack=1, source_image="a.jpg"),
            SheetRecognition("001", layer=1, position=1, stack=1, source_image="a.jpg"),
            SheetRecognition("010", layer=2, position=1, stack=1, source_image="b.jpg"),  # 重复
        ]
    )
    # 数值排序：第10层排在第2层之后，而不是字典序排到前面；
    # 第2层第1位的 010（b.jpg）与 a.jpg 重复被去重。
    assert [item.sheet_no for item in rows] == ["001", "002", "010", "010"]
    assert len(rows) == 4


@pytest.mark.asyncio
async def test_structure_rows_falls_back_without_credentials() -> None:
    rows = await structure_rows(
        credentials=None,
        preferred_model_id="auto",
        raw_rows=[
            SheetRecognition("002", layer=1, position=3, stack=1, source_image="a.jpg"),
            SheetRecognition("001", layer=1, position=1, stack=1, source_image="a.jpg"),
        ],
    )
    assert [item.sheet_no for item in rows] == ["001", "002"]  # 确定性排序


@pytest.mark.asyncio
async def test_structure_rows_uses_llm_when_available(monkeypatch) -> None:
    from backend.services.image import structuring

    captured: dict[str, object] = {}

    async def fake_complete(*, preferred_model_id, credentials, messages, **kwargs):
        captured["model"] = preferred_model_id
        return (
            '[{"sheetNo":"003","layer":1,"position":2,"stack":1,"sourceImage":"a.jpg","note":""}]',
            object(),
            "fake-model",
        )

    monkeypatch.setattr(structuring.GATEWAY, "complete", fake_complete)
    credentials = LlmCredentials(values={"qwen": "k"})
    rows = await structure_rows(
        credentials=credentials,
        preferred_model_id="qwen-test",
        raw_rows=[
            SheetRecognition("003", layer=1, position=2, stack=1, source_image="a.jpg"),
        ],
    )
    assert captured["model"] == "qwen-test"
    assert len(rows) == 1
    assert rows[0].sheet_no == "003"


@pytest.mark.asyncio
async def test_structure_rows_keeps_uncertain_placeholders(monkeypatch) -> None:
    """LLM 整理后，编号为空的占位行（备注无法辨认）不能被丢掉。"""

    from backend.services.image import structuring

    async def fake_complete(*, preferred_model_id, credentials, messages, **kwargs):
        return (
            '[{"sheetNo":"","layer":1,"position":2,"stack":1,"sourceImage":"a.jpg",'
            '"note":"编号无法辨认"}]',
            object(),
            "fake-model",
        )

    monkeypatch.setattr(structuring.GATEWAY, "complete", fake_complete)
    credentials = LlmCredentials(values={"qwen": "k"})
    rows = await structure_rows(
        credentials=credentials,
        preferred_model_id="qwen-test",
        raw_rows=[
            SheetRecognition(
                "",
                layer=1,
                position=2,
                stack=1,
                source_image="a.jpg",
                note="编号无法辨认",
            )
        ],
    )
    assert len(rows) == 1
    assert rows[0].sheet_no == ""
    assert rows[0].note == "编号无法辨认"
    assert rows[0].position == 2


@pytest.mark.asyncio
async def test_structure_rows_falls_back_when_llm_drops_coords(monkeypatch) -> None:
    """LLM 整理结果丢失坐标（例如删掉空位行）时回退确定性排序，保证布局不丢。"""

    from backend.services.image import structuring

    async def fake_complete(*, preferred_model_id, credentials, messages, **kwargs):
        return (
            '[{"sheetNo":"003","layer":1,"position":1,"stack":1,'
            '"sourceImage":"a.jpg","note":""}]',
            object(),
            "fake-model",
        )

    monkeypatch.setattr(structuring.GATEWAY, "complete", fake_complete)
    credentials = LlmCredentials(values={"qwen": "k"})
    rows = await structure_rows(
        credentials=credentials,
        preferred_model_id="qwen-test",
        raw_rows=[
            SheetRecognition("003", layer=1, position=1, stack=1, source_image="a.jpg"),
            SheetRecognition(
                "",
                layer=1,
                position=2,
                stack=1,
                source_image="a.jpg",
                note="空货位",
            ),
        ],
    )
    assert [(item.sheet_no, item.position, item.note) for item in rows] == [
        ("003", 1, ""),
        ("", 2, "空货位"),
    ]


# ---------- excel ----------


def test_excel_filename_safety() -> None:
    assert is_safe_excel_filename("货架图纸识别_20260816_120000.xlsx")
    assert not is_safe_excel_filename("../evil.xlsx")
    assert not is_safe_excel_filename("a/b.xlsx")


def test_clean_cell_text_strips_illegal_xml_characters() -> None:
    """GLM 识别文本夹杂的控制字符会让 xlsx 损坏，必须清洗。"""

    from backend.services.image.excel import clean_cell_text

    dirty = "编号\x00无法辨认\x08\n换行\x1f"
    cleaned = clean_cell_text(dirty)
    assert "\x00" not in cleaned
    assert "\x08" not in cleaned
    assert "\x1f" not in cleaned
    assert "编号无法辨认 换行" == cleaned


def test_write_recognition_excel_roundtrip(tmp_path: Path) -> None:
    """矩阵布局：层标题、货位表头、排行、编号与留空格子。"""

    from openpyxl import load_workbook

    target = write_recognition_excel(
        directory=tmp_path,
        rows=[
            # 第1层：2 排（叠放）× 3 位，缺失格子留空
            SheetRecognition("003\x00", layer=1, position=2, stack=1, source_image="shelf_a.jpg"),
            SheetRecognition("005", layer=1, position=1, stack=2, source_image="shelf_a.jpg"),
        ],
        failures=[ImageRecognitionFailure("shelf_b.jpg", "视觉识别失败：mock")],
        summary="共识别 2 个图纸编号",
    )
    assert target.is_file()
    workbook = load_workbook(target)
    assert workbook.sheetnames == ["图纸编号", "识别失败清单", "识别总结"]
    sheet = workbook["图纸编号"]
    # 行 1：第1层标题
    assert sheet["A1"].value == "第1层"
    # 行 2：货位表头（数据最大 position=2 → 只有第1位/第2位）
    assert sheet["A2"].value == "货位"
    assert sheet["B2"].value == "第1位"
    assert sheet["C2"].value == "第2位"
    # 行 3：排1（003 在 第2位）
    assert sheet["A3"].value == "排1"
    assert sheet["C3"].value == "003"  # 清洗掉控制字符
    assert sheet["B3"].value is None  # 第1位留空
    # 行 4：排2（005 在 第1位）
    assert sheet["A4"].value == "排2"
    assert sheet["B4"].value == "005"
    failure_sheet = workbook["识别失败清单"]
    assert failure_sheet["A2"].value == "shelf_b.jpg"
    assert workbook["识别总结"]["A1"].value == "共识别 2 个图纸编号"


def test_write_recognition_excel_no_rows_adds_notice(tmp_path: Path) -> None:
    """无识别结果时第一个工作表给出说明，避免打开只见空表。"""

    from openpyxl import load_workbook

    target = write_recognition_excel(
        directory=tmp_path,
        rows=[],
        failures=[ImageRecognitionFailure("shelf_b.jpg", "视觉识别失败：mock")],
        summary="未识别到图纸编号",
    )
    workbook = load_workbook(target)
    assert workbook.sheetnames == ["识别结果", "识别失败清单", "识别总结"]
    notice = workbook["识别结果"]
    assert notice["A1"].value == "未识别到图纸编号。"


def test_write_recognition_excel_marks_unknown_cells(tmp_path: Path) -> None:
    """占位行（编号为空）在矩阵里写"空"，位置可见而不是凭空消失。"""

    from openpyxl import load_workbook

    target = write_recognition_excel(
        directory=tmp_path,
        rows=[
            SheetRecognition("", layer=1, position=2, stack=1, source_image="shelf_a.jpg"),
            SheetRecognition("003", layer=1, position=1, stack=1, source_image="shelf_a.jpg"),
        ],
        failures=[],
        summary="识别到 1 个图纸编号，1 处无法辨认",
    )
    workbook = load_workbook(target)
    sheet = workbook["图纸编号"]
    assert sheet["B3"].value == "003"
    assert sheet["C3"].value == "空"


def test_rows_from_boxes_keeps_unknown_placeholder() -> None:
    """标签路径：编号无法辨认的定位框保留为占位行，坐标不丢失。"""

    from backend.services.image.locate import TagBox
    from backend.services.image.service import _rows_from_boxes

    rows, failures = _rows_from_boxes(
        boxes=[TagBox(x=0, y=0, w=10, h=10, row=1, col=2, rank=2)],
        numbers=[None],
        source_image="a.jpg",
    )
    assert len(rows) == 1
    assert rows[0].sheet_no == ""
    assert rows[0].layer == 1
    assert rows[0].position == 2
    assert rows[0].stack == 2
    assert rows[0].note == "编号无法辨认"
    assert len(failures) == 1
    assert failures[0].kind == "quality"


def test_build_excel_filename_format() -> None:
    name = build_excel_filename()
    assert name.startswith("货架图纸识别_")
    assert name.endswith(".xlsx")


# ---------- service SSE ----------


class FakeNoFailClient:
    """固定返回一条识别记录，用于验证 SSE 帧序列。"""

    @property
    def settings(self) -> GLM46VSettings:
        return GLM46VSettings(api_key="fake", endpoint="https://example.test")

    async def analyze_images(self, images, *, prompt: str, max_tokens: int = 6144):
        return {
            "model": "glm-4.6v-flash",
            "content": "003|第1层|第2位|",
            "usage": {},
        }


@pytest.mark.asyncio
async def test_stream_image_recognition_emits_sse_frames(monkeypatch, tmp_path) -> None:
    from backend.services.image import service as image_service

    def fake_settings(_credentials=None):
        return SimpleNamespace(max_image_mb=20)

    monkeypatch.setattr(
        image_service,
        "GLM46VSettings",
        type("FakeSettings", (), {"from_credentials": staticmethod(fake_settings)}),
    )

    def fake_preprocess(*, name, mime_type, data, max_image_mb):
        return SimpleNamespace(name=name, data=data)

    async def fake_segments(*, image, client, source_name):
        return (
            [SheetRecognition("003", layer=1, position=2, stack=1, source_image=source_name)],
            "第1段识别 1 个编号",
            None,
        )

    async def fake_structure(*, credentials, preferred_model_id, raw_rows):
        return raw_rows

    def fake_write(*, directory, rows, failures, summary):
        return tmp_path / "货架图纸识别_test.xlsx"

    monkeypatch.setattr(image_service, "GLM46VClient", lambda settings: FakeNoFailClient())
    monkeypatch.setattr(image_service, "preprocess_image", fake_preprocess)
    monkeypatch.setattr(image_service, "recognize_image_segments", fake_segments)
    monkeypatch.setattr(image_service, "structure_rows", fake_structure)
    monkeypatch.setattr(image_service, "write_recognition_excel", fake_write)

    payload = SimpleNamespace(
        attachments=[
            SimpleNamespace(
                name="shelf_a.jpg",
                mime_type="image/jpeg",
                data="data:image/jpeg;base64,AAAA",
            )
        ]
    )
    frames: list[str] = []
    async for frame in image_service.stream_image_recognition(
        body=payload,
        credentials=None,
        preferred_model_id="auto",
        session_id="sess-1",
    ):
        frames.append(frame)

    joined = "\n".join(frames)
    assert "IMAGE_RESULT" in joined
    assert "TEXT" in joined
    assert "done" in joined
    assert "003" in joined


@pytest.mark.asyncio
async def test_stream_image_recognition_handles_missing_images() -> None:
    from backend.services.image import service as image_service

    payload = SimpleNamespace(attachments=[])
    frames: list[str] = []
    async for frame in image_service.stream_image_recognition(
        body=payload,
        credentials=None,
        preferred_model_id="auto",
        session_id="sess-1",
    ):
        frames.append(frame)
    assert any("没有收到图片附件" in frame for frame in frames)
