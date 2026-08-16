"""把识别结果导出为 openpyxl 生成的 Excel 货架矩阵。

主工作表按货架网格布局：每个层一个区块，区块内每行是一个叠放排、每列是一个
货位，格子为编号或留空（看不清/不确定的位置留空）。识别失败的图片单独记录在
“识别失败清单”工作表。

所有写入单元格的文本都会先清洗 XML 非法控制字符（GLM 识别文本可能夹杂
\\x00-\\x08 等字符，直接写入会让 xlsx 损坏、WPS/Excel 打不开）。
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl.utils import get_column_letter

from backend.services.image.models import ImageRecognitionFailure, SheetRecognition

EXCEL_FILENAME_PATTERN = r"^[\w.\-\u4e00-\u9fff]{1,160}$"
EXCEL_SHEET_PREFIX = "货架图纸识别_"

# XML 1.0 非法控制字符（xlsx 是 zip+XML，含这些字符文件会损坏）。
# 范围外的 \ufffe/\uffff 也是 XML 非字符，一并剥离。
_XML_ILLEGAL_RE = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f\ufffe\uffff]",
)


def clean_cell_text(value: object) -> str:
    """清洗单元格文本：剥离 XML 非法控制字符并规范化空白。"""

    if value is None:
        return ""
    text = str(value)
    cleaned = _XML_ILLEGAL_RE.sub("", text)
    return " ".join(cleaned.split())


def build_excel_filename() -> str:
    """生成 Excel 文件名（含时间戳，路径安全）。"""

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{EXCEL_SHEET_PREFIX}{stamp}.xlsx"


def is_safe_excel_filename(value: str) -> bool:
    """校验下载文件名白名单，防止路径穿越。"""

    return bool(re.match(EXCEL_FILENAME_PATTERN, value))


def write_recognition_excel(
    *,
    directory: Path,
    rows: list[SheetRecognition],
    failures: list[ImageRecognitionFailure],
    summary: str,
) -> Path:
    """生成 Excel 文件并返回其路径。

    主工作表按货架网格布局：每个层一个区块，区块内每一行是一个叠放排，
    每一列是一个货位，格子为编号或留空（看不清/不确定的位置留空）。
    """

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover - 构建环境必带 openpyxl
        raise RuntimeError(
            "图片识别导出 Excel 依赖缺失（"
            f"{exc}）。开发环境请执行 pip install -r requirements.txt；"
            "桌面版需要重新构建（pnpm electron:make:win）后才会包含 openpyxl。"
        ) from exc

    directory.mkdir(parents=True, exist_ok=True)
    filename = build_excel_filename()
    target = directory / filename
    workbook = Workbook()

    if rows:
        from backend.services.image.models import build_layers

        sheet = workbook.active
        sheet.title = "图纸编号"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        board_font = Font(bold=True, color="FFFFFF")
        board_fill = PatternFill("solid", fgColor="70AD47")
        cursor = 1
        for layer_block in build_layers(rows):
            layer = int(layer_block["layer"])
            max_stack = int(layer_block["maxStack"])
            max_position = int(layer_block["maxPosition"])
            cells = layer_block["cells"]  # list[list[dict | None]]  [stack][position]

            # 层标题
            board = sheet.cell(row=cursor, column=1, value=f"第{layer}层")
            board.font = board_font
            board.fill = board_fill
            cursor += 1

            # 表头：货位
            sheet.cell(row=cursor, column=1, value="货位")
            for position in range(1, max_position + 1):
                header = sheet.cell(row=cursor, column=1 + position, value=f"第{position}位")
                header.font = header_font
                header.fill = header_fill
            cursor += 1

            # 每排一行：格子为编号或留空（看不清/不确定）
            for stack_index in range(max_stack):
                stack_label = sheet.cell(row=cursor, column=1, value=f"排{stack_index + 1}")
                stack_label.font = header_font
                stack_label.fill = header_fill
                row_cells = cells[stack_index] if stack_index < len(cells) else []
                for position in range(1, max_position + 1):
                    value = None
                    if position - 1 < len(row_cells):
                        cell = row_cells[position - 1]
                        if cell is not None:
                            value = clean_cell_text(cell["sheetNo"])
                    if value is not None:
                        sheet.cell(row=cursor, column=1 + position, value=value)
                cursor += 1
            cursor += 1  # 层之间的空行分隔

        for column in range(1, 40):
            sheet.column_dimensions[get_column_letter(column)].width = 10
        sheet.freeze_panes = "A2"
    else:
        # 无识别结果时第一个工作表给明确说明，避免打开只见空表头。
        notice = workbook.active
        notice.title = "识别结果"
        notice.cell(row=1, column=1, value="未识别到图纸编号。")
        notice.cell(
            row=2,
            column=1,
            value="请查看“识别失败清单”与“识别总结”；若为免费模型限流，稍等 1-2 分钟后重试。",
        )
        notice.column_dimensions["A"].width = 70

    if failures:
        failure_sheet = workbook.create_sheet("识别失败清单")
        failure_sheet.append(["来源照片", "失败类型", "失败原因"])
        for item in failures:
            failure_sheet.append(
                [
                    clean_cell_text(item.image_name),
                    {
                        "rate_limited": "限流（稍后重试）",
                        "other": "其他",
                        "quality": "照片质量",
                    }.get(item.kind, item.kind),
                    clean_cell_text(item.reason),
                ]
            )
        failure_sheet.column_dimensions["A"].width = 28
        failure_sheet.column_dimensions["B"].width = 20
        failure_sheet.column_dimensions["C"].width = 48

    if summary:
        summary_sheet = workbook.create_sheet("识别总结")
        summary_sheet.cell(row=1, column=1, value=clean_cell_text(summary))
        summary_sheet.column_dimensions["A"].width = 80

    workbook.save(target)
    return target


__all__ = [
    "build_excel_filename",
    "clean_cell_text",
    "is_safe_excel_filename",
    "write_recognition_excel",
]
